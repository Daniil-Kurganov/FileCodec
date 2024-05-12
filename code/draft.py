import openpyxl
from HammingCodec import hamming_codec_encode, hamming_codec_decode

file_source = openpyxl.load_workbook('C:/Users/User/PythonProjects/FileCodec/files/ExcelFileEncode.xlsx', read_only = True)
file_result = openpyxl.Workbook()
bool_operation_is_encode = False
int_count
for sheet_current_source in file_source.worksheets:
    sheet_current_result = file_result.create_sheet(title = sheet_current_source.title)
    for row_current_source in sheet_current_source.iter_rows():
        for cell_current_source in row_current_source:
            if cell_current_source.value is not None:
                if bool_operation_is_encode:
                    sheet_current_result.cell(row = cell_current_source.row, column = cell_current_source.column,
                                              value = hamming_codec_encode(str(cell_current_source.value)))
                else:
                    sheet_current_result.cell(row = cell_current_source.row, column = cell_current_source.column,
                                              value = hamming_codec_decode(str(cell_current_source.value)))
file_source.close()
file_result.remove(file_result['Sheet'])
file_result.save('C:/Users/User/PythonProjects/FileCodec/files/ExcelFileDecode.xlsx')
